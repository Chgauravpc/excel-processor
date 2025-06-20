import json
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import tempfile
import base64
import logging
import os
import sys

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def handler(event, context):
    logger.info("Received event")
    try:
        # Read raw request body
        if not hasattr(event, 'body') or not event.body:
            logger.error("No request body")
            return {
                "statusCode": 400,
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"error": "No request body"})
            }

        # Parse JSON body
        try:
            body = json.loads(event.body.decode('utf-8') if isinstance(event.body, bytes) else event.body)
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON: {str(e)}")
            return {
                "statusCode": 400,
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"error": "Invalid JSON payload"})
            }

        # Validate payload
        if "file" not in body or "filename" not in body:
            logger.error("Missing file or filename")
            return {
                "statusCode": 400,
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"error": "Missing file or filename in payload"})
            }

        filename = body["filename"]
        if not filename.endswith('.xlsx'):
            logger.error("Invalid file extension")
            return {
                "statusCode": 400,
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"error": "File must be .xlsx"})
            }

        # Decode base64 file content
        try:
            file_content = base64.b64decode(body["file"])
        except base64.binascii.Error as e:
            logger.error(f"Invalid base64: {str(e)}")
            return {
                "statusCode": 400,
                "headers": {"Content-Type": "application/json"},
                "body": json.dumps({"error": "Invalid base64 file content"})
            }
        logger.info(f"Decoded file: {filename}")

        # Save to /tmp
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, dir='/tmp') as temp_input:
            temp_input.write(file_content)
            temp_input_path = temp_input.name
            logger.info(f"Saved file to: {temp_input_path}")

        # Process the file
        workbook = openpyxl.load_workbook(temp_input_path)
        worksheet = workbook.active
        max_row = worksheet.max_row
        logger.info(f"Max row: {max_row}")

        # Set row heights
        for row in range(2, 8):
            worksheet.row_dimensions[row].height = 15
        for row in range(8, 13):
            worksheet.row_dimensions[row].height = 0
            worksheet.row_dimensions[row].hidden = True
        for row in range(14, max_row - 1):
            worksheet.row_dimensions[row].height = 27

        # Set column widths
        column_widths = {
            'B': 3, 'E': 9.8, 'F': 13, 'G': 26.3, 'I': 6, 'J': 0,
            'L': 3.2, 'M': 2.7, 'N': 3, 'Q': 2.7, 'U': 7.3, 'V': 5, 'W': 5.5
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
            if width == 0:
                worksheet.column_dimensions[col].hidden = True
        columns_to_hide = ['C', 'P', 'R', 'S', 'T', 'X', 'Y']
        for col in columns_to_hide:
            worksheet.column_dimensions[col].width = 0
            worksheet.column_dimensions[col].hidden = True

        # Define styles
        border_style = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        font_style = Font(size=10)
        font_style_max_row = Font(size=16, bold=True)

        # Format B13:W(max_row - 2)
        for row in range(13, max_row - 1):
            for col_idx in range(2, 24):
                col_letter = get_column_letter(col_idx)
                cell = worksheet[f'{col_letter}{row}']
                cell.border = border_style
                cell.font = font_style
                if col_idx >= 5 and row >= 14:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                if col_idx == 2 and row >= 14:
                    cell.value = row - 13

        # Merge K(max_row - 1):W(max_row - 1)
        if max_row >= 2:
            merge_row = max_row - 1
            worksheet.merge_cells(f'K{merge_row}:W{merge_row}')
            merged_cell = worksheet[f'K{merge_row}']
            merged_cell.font = font_style
            merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            merged_cell.border = border_style
            if merge_row < 14:
                worksheet.row_dimensions[merge_row].height = 27

        # Apply font size 16 and bold to row max_row
        if max_row >= 1:
            for col_idx in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                cell = worksheet[f'{col_letter}{max_row}']
                cell.font = font_style_max_row
                if max_row < 14:
                    worksheet.row_dimensions[max_row].height = 27

        # Save processed file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, dir='/tmp') as temp_output:
            workbook.save(temp_output.name)
            temp_output_path = temp_output.name
            logger.info(f"Saved processed file to: {temp_output_path}")

        # Encode processed file as base64
        with open(temp_output_path, 'rb') as f:
            processed_content = base64.b64encode(f.read()).decode('utf-8')

        # Clean up
        os.unlink(temp_input_path)
        os.unlink(temp_output_path)

        return {
            "statusCode": 200,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({
                "filename": f"processed_{filename}",
                "content": processed_content
            })
        }
    except Exception as e:
        logger.error(f"Server error: {str(e)}")
        return {
            "statusCode": 500,
            "headers": {"Content-Type": "application/json"},
            "body": json.dumps({"error": f"Server error: {str(e)}"})
        }